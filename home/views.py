from django.shortcuts import render

# Create your views here.
from rest_framework.views import APIView
from rest_framework.decorators import api_view
from django.views.decorators.csrf import csrf_exempt
from rest_framework.response import Response
import pandas as pd
import json
from docx import Document
from openpyxl import Workbook, load_workbook
from restapi.custom_log import setup_logger
import logging
setup_logger('log','home')
logobj = logging.getLogger("log")
logobj.info("yes")

class MyNewClass(APIView):
    @csrf_exempt
    @api_view(('POST',))
    def excelmapping(request):
        f = request.FILES['rahul']
        df = pd.read_excel(f,sheet_name='Sheet1',skiprows=1)
        df = df.dropna(thresh=1)
        df = df.fillna(method='ffill',axis='rows')
        df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
        json_val = df.to_json(orient='records')
        json_list = json.loads(json_val)
        return Response(json_list)


    @csrf_exempt
    @api_view(('POST',))
    def docmapping(request):
        f = request.FILES['doc']
        doc = Document(f.file)
        for para in doc.paragraphs:
            if para.text:
                print(para.text)
        return Response("document")

    @csrf_exempt
    @api_view(('POST',))
    def writeexcel(request):
        f = request.FILES['new']
        wb = load_workbook(f.file)
        for sheet in wb.sheetnames:
            for row in wb[sheet].iter_rows():
                for col in row:
                    #col.fill.fgColor.rgb == "" # color 

                    if col.value == "RaJ":
                        col.value = "I love you"
        wb.save(filename='rahul.xlsx')
        return Response("document")


